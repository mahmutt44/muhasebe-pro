from flask import Blueprint, render_template, request, session, jsonify, Response
from flask_login import login_required, current_user
from functools import wraps
from sqlalchemy import func, extract, and_
from datetime import datetime, date
from decimal import Decimal
import calendar
from io import BytesIO

from models import db, Transaction, Product, Customer, CustomerTransaction, Receipt, ReceiptItem, Company

reports = Blueprint('reports', __name__, url_prefix='/reports')


def company_required(f):
    """Company admin ve observer yetkisi kontrolü"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.is_platform_admin:
            return f(*args, **kwargs)
        if not current_user.company_id:
            from flask import flash, redirect, url_for
            flash('Bu işlem için şirket yetkisi gerekiyor.', 'warning')
            return redirect(url_for('main.index'))
        session['company_id'] = current_user.company_id
        return f(*args, **kwargs)
    return decorated_function


def get_company_id():
    """Mevcut şirket ID'sini al - multi-tenant güvenliği"""
    if current_user.is_platform_admin:
        return session.get('selected_company_id')
    return current_user.company_id


# =============================================================================
# RAPOR ANA SAYFASI
# =============================================================================
@reports.route('/')
@company_required
def reports_index():
    """Raporlar ana sayfası - rapor seçimi"""
    company_id = get_company_id()
    if not company_id:
        from flask import flash, redirect, url_for
        flash('Lütfen önce bir işletme seçin.', 'warning')
        return redirect(url_for('main.index'))
    return render_template('reports/index.html')


# =============================================================================
# 1. AYLIK KAR/ZARAR RAPORU
# =============================================================================
@reports.route('/profit-loss')
@company_required
def profit_loss_report():
    """Aylık Kar/Zarar Raporu"""
    company_id = get_company_id()
    
    # Yıl ve ay parametrelerini al
    year = request.args.get('year', type=int, default=datetime.now().year)
    month = request.args.get('month', type=int, default=datetime.now().month)
    
    # Tarih aralığı
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    # Gelir toplamı
    income_total = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.company_id == company_id,
        Transaction.type == 'income',
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).scalar() or 0
    
    # Gider toplamı
    expense_total = db.session.query(func.sum(Transaction.amount)).filter(
        Transaction.company_id == company_id,
        Transaction.type == 'expense',
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).scalar() or 0
    
    # Net kar
    net_profit = Decimal(str(income_total)) - Decimal(str(expense_total))
    
    # Detaylı işlemler
    income_transactions = Transaction.query.filter(
        Transaction.company_id == company_id,
        Transaction.type == 'income',
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).order_by(Transaction.date.desc()).all()
    
    expense_transactions = Transaction.query.filter(
        Transaction.company_id == company_id,
        Transaction.type == 'expense',
        Transaction.date >= start_date,
        Transaction.date <= end_date
    ).order_by(Transaction.date.desc()).all()
    
    # Aylık trend verisi (son 12 ay) - Optimized
    # Tek sorguda tüm aylık verileri al (24 sorgu yerine 1 sorgu)
    monthly_results = db.session.query(
        extract('year', Transaction.date).label('year'),
        extract('month', Transaction.date).label('month'),
        func.sum(
            db.case(
                (Transaction.type == 'income', Transaction.amount),
                else_=0
            )
        ).label('income'),
        func.sum(
            db.case(
                (Transaction.type == 'expense', Transaction.amount),
                else_=0
            )
        ).label('expense')
    ).filter(
        Transaction.company_id == company_id,
        Transaction.date >= date(year - 1, month, 1)  # Son 12 ay
    ).group_by(
        extract('year', Transaction.date),
        extract('month', Transaction.date)
    ).order_by(
        extract('year', Transaction.date),
        extract('month', Transaction.date)
    ).all()
    
    # Sonuçları sözlüğe dönüştür
    monthly_dict = {}
    for r in monthly_results:
        key = f"{int(r.month):02d}/{int(r.year)}"
        monthly_dict[key] = {
            'income': float(r.income or 0),
            'expense': float(r.expense or 0)
        }
    
    # Son 12 ay için eksik ayları 0 ile doldur
    monthly_data = []
    for i in range(11, -1, -1):
        check_month = month - i
        check_year = year
        while check_month <= 0:
            check_month += 12
            check_year -= 1
        
        key = f"{check_month:02d}/{check_year}"
        data = monthly_dict.get(key, {'income': 0.0, 'expense': 0.0})
        
        monthly_data.append({
            'month': key,
            'income': data['income'],
            'expense': data['expense'],
            'profit': data['income'] - data['expense']
        })
    
    return render_template('reports/profit_loss.html',
                         year=year,
                         month=month,
                         month_name=calendar.month_name[month],
                         income_total=float(income_total),
                         expense_total=float(expense_total),
                         net_profit=net_profit,
                         income_transactions=income_transactions,
                         expense_transactions=expense_transactions,
                         monthly_data=monthly_data)


# =============================================================================
# 2. EN ÇOK SATILAN ÜRÜNLER
# =============================================================================
@reports.route('/top-products')
@company_required
def top_products_report():
    """En Çok Satılan Ürünler Raporu"""
    company_id = get_company_id()
    
    limit = request.args.get('limit', type=int, default=10)
    
    # Fiş kalemlerinden ürün satışlarını hesapla
    results = db.session.query(
        Product.id,
        Product.name,
        Product.unit,
        func.sum(ReceiptItem.quantity).label('total_quantity'),
        func.sum(ReceiptItem.total_price).label('total_revenue')
    ).join(
        ReceiptItem, Product.id == ReceiptItem.product_id
    ).join(
        Receipt, ReceiptItem.receipt_id == Receipt.id
    ).filter(
        Receipt.company_id == company_id
    ).group_by(
        Product.id, Product.name, Product.unit
    ).order_by(
        func.sum(ReceiptItem.quantity).desc()
    ).limit(limit).all()
    
    top_products = []
    for r in results:
        top_products.append({
            'id': r.id,
            'name': r.name,
            'unit': r.unit,
            'total_quantity': float(r.total_quantity or 0),
            'total_revenue': float(r.total_revenue or 0)
        })
    
    # Toplam satış istatistikleri
    total_sales = db.session.query(func.sum(ReceiptItem.quantity)).join(
        Receipt, ReceiptItem.receipt_id == Receipt.id
    ).filter(Receipt.company_id == company_id).scalar() or 0
    
    total_revenue = db.session.query(func.sum(ReceiptItem.total_price)).join(
        Receipt, ReceiptItem.receipt_id == Receipt.id
    ).filter(Receipt.company_id == company_id).scalar() or 0
    
    # Kategoriye göre dağılım (unit bazında)
    unit_distribution_results = db.session.query(
        Product.unit,
        func.sum(ReceiptItem.quantity).label('quantity')
    ).join(
        ReceiptItem, Product.id == ReceiptItem.product_id
    ).join(
        Receipt, ReceiptItem.receipt_id == Receipt.id
    ).filter(
        Receipt.company_id == company_id
    ).group_by(Product.unit).all()
    
    # JSON serileştirilebilir formata çevir
    unit_distribution = []
    for row in unit_distribution_results:
        unit_distribution.append({
            'unit': row.unit,
            'quantity': float(row.quantity or 0)
        })
    
    return render_template('reports/top_products.html',
                         top_products=top_products,
                         total_sales=float(total_sales),
                         total_revenue=float(total_revenue),
                         unit_distribution=unit_distribution,
                         limit=limit)


# =============================================================================
# 3. MÜŞTERİ BORÇ LİSTESİ
# =============================================================================
@reports.route('/customer-debts')
@company_required
def customer_debts_report():
    """Müşteri Borç Listesi Raporu - Optimized (N+1 sorunu çözüldü)"""
    company_id = get_company_id()
    
    # Optimizasyon: Tek sorguda tüm müşteri borçlarını hesapla (N+1 sorunu çözüldü)
    # Case when ile borç ve ödemeleri ayrı kolonlarda topla
    results = db.session.query(
        Customer.id,
        Customer.name,
        Customer.phone,
        func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'debt', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        ).label('total_debt'),
        func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'payment', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        ).label('total_paid')
    ).outerjoin(
        CustomerTransaction, Customer.id == CustomerTransaction.customer_id
    ).filter(
        Customer.company_id == company_id
    ).group_by(
        Customer.id, Customer.name, Customer.phone
    ).having(
        # Sadece borcu olan müşteriler (borç > ödeme)
        func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'debt', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        ) > func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'payment', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        )
    ).order_by(
        (func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'debt', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        ) - func.coalesce(
            func.sum(
                db.case(
                    (CustomerTransaction.type == 'payment', CustomerTransaction.amount),
                    else_=0
                )
            ), 0
        )).desc()
    ).all()
    
    # Sonuçları formatla
    customer_debts = []
    total_debt = 0
    total_paid = 0
    
    for r in results:
        balance = float(r.total_debt) - float(r.total_paid)
        customer_debts.append({
            'id': r.id,
            'name': r.name,
            'phone': r.phone,
            'total_debt': float(r.total_debt),
            'total_paid': float(r.total_paid),
            'balance': balance
        })
        total_debt += float(r.total_debt)
        total_paid += float(r.total_paid)
    
    # Borçlu müşteri sayısı
    debtors_count = len(customer_debts)
    
    return render_template('reports/customer_debts.html',
                         customer_debts=customer_debts,
                         total_debt=total_debt,
                         total_paid=total_paid,
                         net_debt=total_debt - total_paid,
                         debtors_count=debtors_count)


# =============================================================================
# 4. STOK RAPORU
# =============================================================================
@reports.route('/stock')
@company_required
def stock_report():
    """Stok Raporu - Optimized (SQL aggregation ile hesaplama)"""
    company_id = get_company_id()
    
    # Düşük stok filtreleme
    show_low_stock = request.args.get('low_stock', type=bool, default=False)
    
    # Optimizasyon: SQL ile toplam stok değeri ve düşük stok sayısını hesapla
    # Aynı zamanda düşük stok filtrelemesi de SQL'de yapılabilir
    base_query = Product.query.filter_by(company_id=company_id)
    
    if show_low_stock:
        # Sadece düşük stoklu ürünleri göster
        products = base_query.filter(
            Product.stock_quantity < Product.stock_threshold
        ).all()
    else:
        # Tüm ürünleri al
        products = base_query.all()
    
    stock_data = []
    total_stock_value = 0
    
    for product in products:
        stock_value = float(product.stock_quantity or 0) * float(product.purchase_price or 0)
        is_low = float(product.stock_quantity or 0) < float(product.stock_threshold or 10)
        
        stock_data.append({
            'id': product.id,
            'name': product.name,
            'unit': product.unit,
            'stock_quantity': float(product.stock_quantity or 0),
            'stock_threshold': float(product.stock_threshold or 10),
            'purchase_price': float(product.purchase_price or 0),
            'unit_price': float(product.unit_price or 0),
            'stock_value': stock_value,
            'is_low_stock': is_low
        })
        total_stock_value += stock_value
    
    # SQL ile düşük stok sayısını ve toplam değerleri hesapla
    low_stock_count = db.session.query(func.count(Product.id)).filter(
        Product.company_id == company_id,
        Product.stock_quantity < Product.stock_threshold
    ).scalar() or 0
    
    # Stok durumu istatistikleri
    total_products = len(products)
    
    return render_template('reports/stock.html',
                         stock_data=stock_data,
                         total_products=total_products,
                         low_stock_count=low_stock_count,
                         total_stock_value=total_stock_value,
                         show_low_stock=show_low_stock)


# =============================================================================
# EXPORT İŞLEMLERİ
# =============================================================================
@reports.route('/export/<report_type>/<format>')
@company_required
def export_report(report_type, format):
    """Raporları Excel veya PDF olarak dışa aktar"""
    company_id = get_company_id()
    
    if format == 'excel':
        return export_excel(report_type, company_id)
    elif format == 'pdf':
        return export_pdf(report_type, company_id)
    else:
        return jsonify({'error': 'Geçersiz format'}), 400


def export_excel(report_type, company_id):
    """Excel formatında dışa aktar"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        
        if report_type == 'profit_loss':
            ws.title = "Kar Zarar"
            ws.append(['Tarih', 'Tür', 'Tutar', 'Açıklama'])
            transactions = Transaction.query.filter_by(company_id=company_id).order_by(Transaction.date.desc()).all()
            for t in transactions:
                ws.append([t.date, 'Gelir' if t.type == 'income' else 'Gider', float(t.amount), t.description])
        
        elif report_type == 'stock':
            ws.title = "Stok Raporu"
            ws.append(['Ürün', 'Birim', 'Stok', 'Alış Fiyatı', 'Stok Değeri'])
            products = Product.query.filter_by(company_id=company_id).all()
            for p in products:
                stock_value = float(p.stock_quantity or 0) * float(p.purchase_price or 0)
                ws.append([p.name, p.unit, float(p.stock_quantity or 0), float(p.purchase_price or 0), stock_value])
        
        elif report_type == 'top_products':
            ws.title = "En Çok Satan Ürünler"
            ws.append(['Ürün', 'Birim', 'Satış Miktarı', 'Toplam Gelir'])
            results = db.session.query(
                Product.name,
                Product.unit,
                func.sum(ReceiptItem.quantity).label('total_quantity'),
                func.sum(ReceiptItem.total_price).label('total_revenue')
            ).join(
                ReceiptItem, Product.id == ReceiptItem.product_id
            ).join(
                Receipt, ReceiptItem.receipt_id == Receipt.id
            ).filter(
                Receipt.company_id == company_id
            ).group_by(
                Product.id, Product.name, Product.unit
            ).order_by(
                func.sum(ReceiptItem.quantity).desc()
            ).limit(10).all()
            for r in results:
                ws.append([r.name, r.unit, float(r.total_quantity or 0), float(r.total_revenue or 0)])
        
        elif report_type == 'customer_debts':
            ws.title = "Müşteri Borçları"
            ws.append(['Müşteri', 'Telefon', 'Toplam Borç', 'Ödenen', 'Kalan Borç'])
            customers = Customer.query.filter_by(company_id=company_id).all()
            for customer in customers:
                debts = db.session.query(func.sum(CustomerTransaction.amount)).filter(
                    CustomerTransaction.customer_id == customer.id,
                    CustomerTransaction.type == 'debt'
                ).scalar() or 0
                payments = db.session.query(func.sum(CustomerTransaction.amount)).filter(
                    CustomerTransaction.customer_id == customer.id,
                    CustomerTransaction.type == 'payment'
                ).scalar() or 0
                balance = float(debts) - float(payments)
                if balance > 0:
                    ws.append([customer.name, customer.phone or '', float(debts), float(payments), balance])
        
        # Sütun genişlikleri
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Başlık stili
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{report_type}_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    
    except ImportError:
        return jsonify({'error': 'openpyxl kütüphanesi kurulu değil'}), 500


def export_pdf(report_type, company_id):
    """PDF formatında dışa aktar"""
    return jsonify({'message': 'PDF export yakında eklenecek'}), 200
